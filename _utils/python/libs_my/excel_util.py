#!python
# -*- coding:utf-8 -*-
"""
公用函数(excel的操作)
Created on 2017/6/23
Updated on 2019/1/18
@author: Holemar

依赖第三方库:
    xlrd>=0.9.0,<1.0
    XlsxWriter==0.7.7
    openpyxl==2.4.8
    comtypes==1.1.4
    xlwings==0.11.7
"""
import os
import types
import time
import datetime
import random
import decimal
import logging

import xlrd
import xlsxwriter
import openpyxl

from .str_util import to_str, to_unicode
from .file_util import download_file, remove

try:
    import cStringIO as StringIO
except:
    import StringIO


__all__=('ExcelExport', 'excel_reader', 'excel_abstract')


def get_object_value(obj, code, default=None):
    """
    获取表格传入对象的值
    :param {object|dict} obj: 要获取值的对象
    :param {string|function} code: 要获取值的属性名、key，或者函数
    :param {任意} default: 获取不到值时的默认值
    :return {任意}: 获取的值
    """
    if isinstance(code, (types.FunctionType, types.MethodType)):
        # 值是函数时，避免报异常
        try:
            value = code(obj)
        except:
            value = default
    elif isinstance(obj, dict):
        value = obj.get(code, default)
    else:
        value = getattr(obj, code, default)
    return value


def try_float(value):
    """
    将value转成float类型
    :param {string|int|long|float|decimal} value: 要转换的值
    :return {float}: 转换后的值
    """
    try:
        return float(value)
    except:
        return 0.0


class ExcelExport(object):
    """
    导出 Excel 的基类
    """

    export_name = None # 导出的文件名(不包含文件后缀名)

    def __init__(self, data=[], export_name=None, file_path=None, **kwargs):
        """
        :param data: 需要导出的数据集
        :param export_name: 导出的文件名(不包含文件后缀名)
        :param file_path: 文件路径(不包含文件名，没有则放内存)
        :param kwargs:额外需要保存的参数
        """
        self.data = data
        self.is_close = False

        # 保存额外参数
        if kwargs:
            for k,v in kwargs.items():
                setattr(self, k, v)

        # 保存，导出的文件名
        if export_name:
            self.export_name = export_name

        # 如果有文件路径，到生成文件到硬盘。没有则生成到内存
        if file_path:
            #self.output = os.path.join(os.getcwd(), self.file_dir or 'static/export', datetime.datetime.now().strftime('%Y%m%d'), self.get_export_name())
            #file_path = os.path.dirname(self.output)
            self.output = os.path.join(file_path, self.get_export_name())
            # 防止目录不存在导致异常
            if not os.path.isdir(file_path):
                os.makedirs(file_path)
        else:
            self.output = StringIO.StringIO() # excel放内存

        self.workbook = xlsxwriter.Workbook(self.output)
        self.add_formats()
        self.create_excel()

    def add_formats(self):
        """
        默认的格式
        """
        self.title_format = self.workbook.add_format({'align':'center', 'valign':'vcenter','font_size':20,'bold':True}) # 顶部大标题
        self.head_format = self.workbook.add_format({'align':'center', 'valign':'vcenter', "bg_color": "#ebf5fa", 'border': 1}) # 表头格式
        self.sum_format = self.workbook.add_format({'align':'center', 'valign':'vcenter', "bg_color": "#ebf5fa", 'border': 1}) # 合计行格式
        self.left_content = self.workbook.add_format({'align':'left', 'valign':'vcenter', 'border': 1}) # 文字格式
        self.center_content = self.workbook.add_format({'align':'center', 'valign':'vcenter', 'border': 1}) # 文字格式居中
        self.percent_format = self.workbook.add_format({'num_format': '0.00%;[Red]-0.00%;_ * -??_ ;_ @_ ', 'border': 1}) # 比例格式
        self.int_format = self.workbook.add_format({'num_format': '_ * #,##0_ ;_ * -#,##0_ ;_ * -??_ ;_ @_ ', 'border': 1}) # 整型格式
        self.float_format = self.workbook.add_format({'num_format': '_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * -??_ ;_ @_ ', 'border': 1})  # 浮点型格式
        self.money_format = self.float_format
        self.date_format = self.workbook.add_format({'num_format': 'yyyy-m-d', 'align': 'left', 'valign':'vcenter', 'border': 1}) # 日期
        self.time_format = self.workbook.add_format({'num_format': 'yyyy-m-d h:mm:ss', 'align': 'left', 'valign':'vcenter', 'border': 1}) # 时间

    def add_sheets(self):
        raise Exception("需要重写此方法，用于生成excel：add_sheets")

    def create_excel(self):
        """
        生成excel内容
        """
        self.add_sheets()
        self.workbook.close()


    def get_export_name(self):
        if not self.export_name:
            raise Exception("需要配置导出文件名称：export_name")
        export_name = self.export_name
        if not export_name.lower().endswith('.xlsx'):
            export_name = u'%s.xlsx' % export_name
        return export_name

    def get_output(self):
        """
        返回生成的excel二进制内容
        """
        self.output.seek(0)
        return self.output.read()

    def close(self):
        if not self.is_close:
            self.workbook.close()
            if isinstance(self.output, StringIO.StringIO):
                self.output.close()
            del self.data
            del self.output
            del self.workbook
            self.is_close = True

    def __del__(self):
        self.close()
        super(ExcelExport, self).__del__()

    def set_options(self, worksheet, **kwargs):
        """
        加入此标签页的配置
        :param worksheet: 标签页
        :param head_row:表头占多少行(默认占2行)
        :param first_row:表格开始行号
        :param first_col:表格开始列号
        :param head_format:表头格式
        :param source:数据源
        :param content_format:内容格式
        :param head_datas:表头内容
        :return:
        """
        worksheet.worksheet_options = kwargs.copy()

    def get_last_row_index(self, worksheet):
        """
        获取此标签页的最后一行数据的下标(仅当 write_table_data 执行完之后生效)
        :param worksheet: 标签页
        :return:
        """
        return getattr(worksheet, 'last_row_index', 0)

    def set_table_col(self, worksheet, options=None, parent_head='', title=None, child_list=None):
        """
        设置各列信息
        :param worksheet: 标签页
        :param options: 表格配置信息。head_row:表头占多少行(默认占2行)， first_row:开始行号，first_col:开始列号，head_format:表头格式，source:数据源，content_format:内容格式，head_datas:表头内容
        :param parent_head: 一级表头的显示文字
        :param title: 标题的批注
        :param child_list: 这一列(或者多列)的内容信息。 code:数据源里的字段名；default:默认值；head:二级表头的显示文字(可以是空字符串)；width:宽度；format:内容样式、格式；type:sum 表示累计，index 表示序号；title:批注。
        """
        options = options or getattr(worksheet, 'worksheet_options', {})
        worksheet.worksheet_options = options
        head_row = options.get('head_row', 2)
        first_row = options.get('first_row', 0)
        first_col = options.get('first_col', 0)
        head_format = options.get('head_format', None) or self.head_format
        content_format = options.get('content_format', None) or self.left_content
        head_datas = options.get('head_datas', [])

        child_list = [child_list] if isinstance(child_list, dict) else child_list
        # 表头只占一行的
        if head_row == 1:
            first_row -= 1
        # 合并一级表头
        else:
            if len(child_list) == 1:
                if child_list[0].get('head'):
                    #worksheet.write(first_row, first_col, parent_head, head_format) # 无合并
                    head_datas.append({'type':'write', "cell" : (first_row, first_col), "content":parent_head, "format":head_format, 'comment':title})
                else:
                    #worksheet.merge_range(first_row, first_col, first_row+1, first_col, parent_head, head_format)
                    head_datas.append({'type':'merge', 'cell' : (first_row, first_col, first_row+1, first_col), "content":parent_head, "format":head_format, 'comment':title})
            else:
                #worksheet.merge_range(first_row, first_col, first_row, first_col + len(child_list) - 1, parent_head, head_format)
                head_datas.append({'type':'merge', 'cell' : (first_row, first_col, first_row, first_col + len(child_list) - 1), "content":parent_head, "format":head_format, 'comment':title})
        for child in child_list:
            # 表格内容
            options.setdefault('codes', []).append(child.get('code'))
            options.setdefault('defaults', []).append(child.get('default', ''))
            width = child.get('width', None)
            if width is None:
                width = len(to_unicode(child.get('head', parent_head))) * 2
                width = 12 if width < 12 else width
            column = {'width': width, "format": child.get('format', content_format)}
            options.setdefault('columns', []).append(column)
            options.setdefault('type', []).append(child.get('type'))
            # 设置列宽
            worksheet.set_column(first_col, first_col, width=width)
            # 二级表头
            #worksheet.write(first_row + 1, first_col, child.get('head', ''), head_format)
            head_datas.append({'type':'write', "cell" : (first_row + 1, first_col), "content":child.get('head', parent_head), "format":head_format, 'comment':child.get('title', title)})
            first_col += 1

        options['first_col'] = first_col
        options['head_datas'] = head_datas

    def write_table_data(self, worksheet, options=None, add_sum_line=True):
        """
        写入表格信息
        :param worksheet: 标签页
        :param options: 表格配置信息。first_row:开始行号；first_col:开始列号；head_format:表头格式；source:数据源；codes:数据源里的字段名；defaults:默认值列表；columns:内容样式、格式；type:sum 表示累计，index 表示序号。
        :param add_sum_line: 是否需要加上统计行
        """
        options = options or getattr(worksheet, 'worksheet_options', {})
        worksheet.worksheet_options = options
        data = []
        head_row = options.get('head_row', 2)
        first_row = options.get('first_row', 0)
        columns = options.get('columns', [])
        codes = options.get('codes', [])
        defaults = options.get('defaults', [])
        col_type = options.get('type', [])
        source = options.get('source', self.data)
        head_datas = options.get('head_datas', [])
        # 各行数据
        col_index = 0
        for obj in source:
            tem = []
            for index in range(len(codes)):
                code = codes[index]
                default = defaults[index]
                this_type = col_type[index]
                # 序号
                if this_type == 'index':
                    col_index += 1
                    tem.append(col_index)
                else:
                    value = get_object_value(obj, code, default)
                    tem.append(value)
            data.append(tem)

        # 不知道什么原因，直接用 add_table 会导致生成的Excel报错，或者没有了筛选功能。故改成下面的遍历 write 写法。
        #worksheet.add_table(first_row+2, 0, len(data)+first_row+1, len(columns)-1, {"data": data, "header_row": False, "columns":columns})
        row_index = first_row + 1 if head_row == 1 else first_row + 2
        for row_data in data:
            for col_index in range(len(codes)):
                format = columns[col_index].get('format')
                value = row_data[col_index]
                if not format:
                    if isinstance(value, (int, long)):
                        format = self.int_format
                    elif isinstance(value, (float, decimal.Decimal)):
                        format = self.float_format
                    elif isinstance(value, (time.struct_time, datetime.datetime)):
                        format = self.time_format
                    elif isinstance(value, datetime.date):
                        format = self.date_format
                    else:
                        format = self.left_content
                worksheet.write(row_index, col_index, value, format)
            row_index += 1

        # 统计行
        if add_sum_line:
            for col_index in range(len(codes)):
                code = codes[col_index]
                this_type = col_type[col_index]
                if col_index == 0:
                    value = u'合计：'
                elif this_type == 'sum':
                    amount = sum([try_float(get_object_value(obj, code, 0)) for obj in source])
                    value = amount
                else:
                    value = u''
                worksheet.write(row_index, col_index, value, self.sum_format)
            row_index += 1

        # 设置最后行坐标
        setattr(worksheet, 'last_row_index', row_index - 1)

        # 写表头
        for head_option in head_datas:
            c_type = head_option.get('type')
            cell = head_option.get('cell')
            content = head_option.get('content')
            format = head_option.get('format') or self.head_format
            title = head_option.get('comment')
            if isinstance(cell, basestring):
                param = (cell, content, format)
            elif isinstance(cell, (list, tuple)):
                param = list(cell) + [content, format]
            else:
                raise RuntimeError(u'传递错误参数！')
            if 'write' == c_type:
                worksheet.write(*param)
            elif 'merge' == c_type:
                worksheet.merge_range(*param)
            if title:
                worksheet.write_comment(cell[0], cell[1], title)


def excel_reader(url, sheet_name=None):
    """
    读取 Excel 内容
    :param {string} url: 需读取的 excel 文件路径
    :param {string} sheet_name: 需读取的 标签页名称
    :return {dict}: Excel 内容的dict, {标签页名称:按行列组成的二维数组table}
    """
    # 文件必须先下载到本地，否则无法读取
    url_lower = url.lower()
    tem_file = False # 是否需要删除临时文件的标识
    if url_lower.startswith('http:') or url_lower.startswith('https:'):
        now = datetime.datetime.now()
        file_path = os.path.join('logs', 'excel', now.strftime('%Y%m%d'), now.strftime('%H%M%S'), str(random.randint(1000, 9999)) + '.xlsx')
        download_file(url, file_path)
        tem_file = True
    # 可能是本地硬盘地址(本机测试时用)
    else:
        file_path = url

    file_path = to_unicode(file_path)
    if not os.path.exists(file_path):
        raise Exception(u"文件不存在，请检查路径是否正确。")

    try:
        # office 2007 文件读取
        return excel_openpyxl_reader(file_path, sheet_name=sheet_name)
    except:
        logging.warning(u'Excel文件读取失败', exc_info=True)

        # office 97-2003 文件读取
        return excel_xlrd_reader(file_path, sheet_name=sheet_name)
    finally:
        # 删除临时下载的excel文件，没必要存
        if tem_file and os.path.isfile(file_path):
            try:
                remove(file_path)
            except: # 不一定有删除权限
                pass


def excel_openpyxl_reader(url, sheet_name=None):
    """
    读取 2007格式的 Excel 内容
    :param {string} url: 需读取的 excel 文件路径
    :param {string} sheet_name: 需读取的 标签页名称
    :return {dict}: Excel 内容的dict, {标签页名称:按行列组成的二维数组table}
    """
    # office 2007 文件读取
    workbook = openpyxl.load_workbook(filename=url, read_only=True, guess_types=True)
    data = {} # 数据容器,内容为 {标签页名称:内容}
    #active_sheet = workbook.active # 被选中的标签页
    # 遍历各标签页
    for sheet in workbook:
        if sheet_name and sheet.title != sheet_name: continue
        max_row = sheet.max_row # 最大行数
        max_column = sheet.max_column # 最大列数
        keep_data = [] # 解析类型后的新结果
        for row_num in range(1, max_row + 1):
            row_values = []
            keep_data.append(row_values)
            for col_num in range(1, max_column + 1):
                value = sheet.cell(row=row_num,column=col_num).value
                row_values.append(value)
        if max_row == 1 and max_column == 1 and keep_data == [[None]]: continue # 没有内容的标签页，不处理
        data[sheet.title] = keep_data
    if sheet_name:
        if data:
            return data.values()[0]
        else:
            return [[]]
    return data


def excel_xlrd_reader(url, sheet_name=None):
    """
    读取 office 97-2003格式的 Excel 内容
    :param {string} url: 需读取的 excel 文件路径
    :param {string} sheet_name: 需读取的 标签页名称
    :return {dict}: Excel 内容的dict, {标签页名称:按行列组成的二维数组table}
    """
    # office 97-2003 文件读取
    try:
        workbook = xlrd.open_workbook(url, encoding_override='utf-8')
    except UnicodeDecodeError, e:
        workbook = xlrd.open_workbook(url, encoding_override='gbk')
    sheets = workbook.sheets()
    data = {} # 数据容器,内容为 {标签页名称:内容}
    # 遍历各标签页
    for sheet in sheets:
        if sheet_name and sheet.name != sheet_name: continue
        # 本标签页的数据容器,内容是一个二维数组,一个按行列组成的table
        table = sheet._cell_values
        if not table: continue # 没有内容的标签页，不处理
        # 类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
        cell_types = sheet._cell_types
        keep_data = [] # 解析类型后的新结果
        for col_values in table:
            row_values = []
            row_types = cell_types[len(keep_data)]
            keep_data.append(row_values)
            for value in col_values:
                _type = row_types[len(row_values)]
                # value的类型转换
                # 0 empty, 2 number
                if _type in (0, 2):
                    pass
                # 1 string (由于兼容GBK格式的Excel，故需要编码转换成utf8)
                elif _type == 1:
                    value = to_str(value)
                # 3 date
                elif _type == 3:
                    value = xlrd.xldate.xldate_as_datetime(value, 0)
                # 4 boolean
                elif _type == 4:
                    value = bool(value)
                row_values.append(value)
        data[sheet.name] = keep_data

    # 返回结果
    if sheet_name:
        if data:
            return data.values()[0]
        else:
            return [[]]
    return data


def _get_excel_value(value):
    """
    读取Excel中的内容
    :param {任意} value: excel中的内容值
    :return {任意}: 转型后的内容值
    """
    if value is None:
        return None
    if isinstance(value, basestring):
        value = value.strip()
        if value.startswith("'"):
            value = value[1:]
        return value
    if isinstance(value, float):
        if int(value) == value:
            return int(value)
    return value


def excel_abstract(sheet_data, title_dict, title_list_dict, need_title):
    """
    excel解析，提取所需的数据并整合成固定的格式
    :param {list<list>} sheet_data: excel标签页内容(仅一页的)，由 excel_reader 返回的各标签页内容
    :param {dict<list>} title_dict: 字段名列表及对应key的单列取值集合，取值value都只取一列，格式是： {'key1':'字段名11', 'key2':['字段名21','字段名22','字段名23']}
    :param {dict<list>} title_list_dict: 字段名列表及对应key的多列取值集合，取值value可以取多列(返回值的列表)，格式是： {'key3':['字段名31','字段名32','字段名33'], 'key4':['字段名41','字段名42','字段名43']}
    :param {list} need_title: 必须的标题key列表，默认全部的 title_dict 都需要， 格式是： ['key1', 'key2', 'key3', 'key4']
    :return {list<dict>}:解析出的数据列表，各内容以dict形式拼接。如： [{'key1':111,'key2':222, 'key3':[11,22,33], 'key4':[44,55,66,]}]
    """
    result = [] # Excel解析结果
    if not sheet_data:
        return result
    if not title_dict or not isinstance(title_dict, dict):
        raise RuntimeError(u'没有取值字段名，无法解析Excel！')
    if not need_title:
        need_title = title_dict.keys()
    title_list_dict = title_list_dict or {}

    index_line = 0 # 取值行标
    read_line_keys = False # 标题列是否已取值成功
    index_dict = {k:None for k in title_dict} # 单列取值集合
    indexes_dict = {k:[] for k in title_list_dict} # 多列取值集合
    need_title_keys = [] # 标题行的列名称列表
    for k, col_names in title_dict.iteritems():
        if isinstance(col_names, (list, tuple, set)):
            col_names.extend([unicode(v).upper() for v in col_names]) # 忽略列的大小写
            if k in need_title:
                need_title_keys.extend(col_names)
        elif isinstance(col_names, basestring):
            col_name_list = [col_names, unicode(col_names).upper()] # 忽略列的大小写
            title_dict[k] = col_name_list
            if k in need_title:
                need_title_keys.extend(col_name_list)
    for k, col_name_list in title_list_dict.iteritems():
        col_name_list.extend([unicode(v).upper() for v in col_name_list]) # 忽略列的大小写
        if k in need_title:
            need_title_keys.extend(col_name_list)

    # 遍历excel表
    for line in sheet_data:
        # 先解析出标题
        if read_line_keys is False:
            line_upper = [(unicode(v).upper() if isinstance(v, basestring) else v) for v in line]
            tem_line = line + line_upper
            if len([l for l in tem_line if l in need_title_keys]) >= 1:
                read_line_keys = True
                # 单列取值
                for k, col_name_list in title_dict.iteritems():
                    for col_name in col_name_list:
                        if col_name in line:
                            index_dict[k] = line.index(col_name)
                            break
                        if col_name in line_upper:
                            index_dict[k] = line_upper.index(col_name)
                            break
                # 多列取值
                for k, col_name_list in title_list_dict.iteritems():
                    for col_name in col_name_list:
                        if col_name in line:
                            indexes_dict[k].append(line.index(col_name))
                        if col_name in line_upper:
                            indexes_dict[k].append(line_upper.index(col_name))
            index_line += 1
        # 读取数据
        else:
            # 获取内容
            data = {}
            result.append(data)
            # 单列取值
            for k, index in index_dict.iteritems():
                if index is None or not isinstance(index, int):continue
                data[k] = _get_excel_value(line[index])
            # 多列取值
            for k, index_list in indexes_dict.iteritems():
                if not index_list:continue
                value_list = data[k] = []
                for index in index_list:
                    value_list.append(_get_excel_value(line[index]))
    return result
